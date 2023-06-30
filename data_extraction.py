import openpyxl
from datetime import datetime

def print_to_excel(file_path, sheet_name, column_number, data_file):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the target sheet
    sheet = workbook[sheet_name]

    # Read data from the text file and convert to integers
    with open(data_file, 'r') as file:
        data = file.readlines()

    data = [int(value.strip()) for value in data]

    # Get the total number of rows and columns in Table3
    max_rows = sheet.max_row
    max_cols = sheet.max_column

    # Find the current day
    current_day = datetime.now().strftime("%A")

    # Find the column index for the matching day
    days = sheet[1]  # Assuming the days are in the first row
    day_column_index = None
    for cell in days:
        if cell.value == current_day:
            day_column_index = cell.column
            break

    if day_column_index is None:
        print(f"No matching day found for '{current_day}' in the table.")
        return

    # Print the data in the specified column of Table3
    for row_num in range(2, max_rows + 1):
        value = data[row_num - 2]

        # Convert value to int
        value = int(value)

        sheet.cell(row=row_num, column=column_number).value = value

        # Check if the day matches the current day and update the corresponding column
        if column_number == day_column_index:
            sheet.cell(row=row_num, column=1).value = current_day

    # Save the changes
    workbook.save(file_path)
    print("Data printed successfully.")

# Provide the file path and sheet name
file_path = "beach_robot_data.xlsx"
sheet_name = "Data"

# Define the mappings of column numbers and data files
column_files_mapping = {
    2: "output2.txt",
    3: "output3.txt",
    4: "output4.txt",
    5: "output5.txt"
}

# Process each column and corresponding data file
for column_number, data_file in column_files_mapping.items():
    # Call the function for each column and data file
    print_to_excel(file_path, sheet_name, column_number, data_file)
